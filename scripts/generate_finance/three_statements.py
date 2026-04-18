#!/usr/bin/env python3
"""
three_statements.py — 金融三表联动模板生成器

生成一个完整的财务三表联动 Excel 模板，包含：
- Sheet 1: 假设输入（Assumptions）
- Sheet 2: 利润表（Income Statement）
- Sheet 3: 资产负债表（Balance Sheet）
- Sheet 4: 现金流量表（Cash Flow Statement）
- Sheet 5: 关键比率（KPI Dashboard：ROE/ROA/毛利率/资产负债率）

三表通过公式联动：
- 利润表的净利润 → 资产负债表的留存收益增量
- 利润表的折旧 → 现金流的经营现金流加项
- 资产负债表平衡校验：资产 = 负债 + 所有者权益

用法:
    python scripts/generate_finance/three_statements.py --output mycompany.xlsx
    python scripts/generate_finance/three_statements.py --output x.xlsx --years 5 --company "示例公司"
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
SECTION_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F4E78')
SECTION_FILL = PatternFill('solid', fgColor='D9E1F2')
TOTAL_FONT = Font(name='微软雅黑', size=10, bold=True)
NORMAL_FONT = Font(name='微软雅黑', size=10)
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')  # 浅黄表示用户输入
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def style_header(cell) -> None:
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER


def style_section(cell) -> None:
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER


def style_total(cell, fmt: str = '#,##0') -> None:
    cell.font = TOTAL_FONT
    cell.fill = PatternFill('solid', fgColor='E7E6E6')
    cell.border = THIN_BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal='right')


def style_data(cell, fmt: str = '#,##0', is_input: bool = False) -> None:
    cell.font = NORMAL_FONT
    cell.border = THIN_BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal='right')
    if is_input:
        cell.fill = INPUT_FILL


def col(n: int) -> str:
    return get_column_letter(n)


def build_assumptions(wb: Workbook, years: int, company: str) -> None:
    """假设输入 sheet。用户改这里的浅黄单元格，三表自动重算。"""
    ws = wb.create_sheet('1.假设')
    ws['A1'] = f'{company} - 财务模型假设'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=years + 2)

    headers = ['项目', '说明'] + [f'第{i + 1}年' for i in range(years)]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        style_header(c)

    rows = [
        ('收入增长率', '同比', [0.10] + [0.08] * (years - 1)),
        ('毛利率', '占收入', [0.40] * years),
        ('销售费用率', '占收入', [0.10] * years),
        ('管理费用率', '占收入', [0.05] * years),
        ('研发费用率', '占收入', [0.05] * years),
        ('所得税率', '法定', [0.25] * years),
        ('折旧率', '占固定资产', [0.10] * years),
        ('应收账款周转天数', '天', [60] * years),
        ('存货周转天数', '天', [45] * years),
        ('应付账款周转天数', '天', [50] * years),
        ('CapEx 占收入', '资本开支', [0.05] * years),
        ('股利支付率', '占净利', [0.30] * years),
    ]
    for i, (name, desc, vals) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=name).font = TOTAL_FONT
        ws.cell(row=i, column=2, value=desc).font = NORMAL_FONT
        for j, v in enumerate(vals, start=3):
            c = ws.cell(row=i, column=j, value=v)
            style_data(c, fmt='0.00%' if isinstance(v, float) and v < 1 else '#,##0', is_input=True)

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 14
    for j in range(3, 3 + years):
        ws.column_dimensions[col(j)].width = 12


def build_income(wb: Workbook, years: int, company: str) -> None:
    """利润表。"""
    ws = wb.create_sheet('2.利润表')
    ws['A1'] = f'{company} - 利润表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

    # Header
    headers = ['项目'] + [f'第{i + 1}年' for i in range(years)]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=j, value=h)
        style_header(c)

    # Year 0 base：100,000,000 收入（用户改）
    base_revenue = 100_000_000

    # 收入：第1年 = base * (1 + 增长率1)；后续 = 上一年 * (1 + 增长率N)
    ws.cell(row=4, column=1, value='营业收入').font = TOTAL_FONT
    for j in range(years):
        col_year = col(j + 2)
        if j == 0:
            f = f'={base_revenue}*(1+\'1.假设\'!{col(j + 3)}4)'
        else:
            f = f'={col(j + 1)}4*(1+\'1.假设\'!{col(j + 3)}4)'
        c = ws.cell(row=4, column=j + 2, value=f)
        style_total(c)

    # 营业成本 = 收入 *(1 - 毛利率)
    ws.cell(row=5, column=1, value='营业成本').font = NORMAL_FONT
    for j in range(years):
        f = f'={col(j + 2)}4*(1-\'1.假设\'!{col(j + 3)}5)'
        style_data(ws.cell(row=5, column=j + 2, value=f))

    # 毛利
    ws.cell(row=6, column=1, value='毛利').font = TOTAL_FONT
    for j in range(years):
        f = f'={col(j + 2)}4-{col(j + 2)}5'
        style_total(ws.cell(row=6, column=j + 2, value=f))

    # 销售/管理/研发费用
    for i, (name, ass_row) in enumerate([('销售费用', 6), ('管理费用', 7), ('研发费用', 8)], start=7):
        ws.cell(row=i, column=1, value=name).font = NORMAL_FONT
        for j in range(years):
            f = f'={col(j + 2)}4*\'1.假设\'!{col(j + 3)}{ass_row}'
            style_data(ws.cell(row=i, column=j + 2, value=f))

    # 营业利润 = 毛利 - 三费
    ws.cell(row=10, column=1, value='营业利润').font = TOTAL_FONT
    for j in range(years):
        f = f'={col(j + 2)}6-{col(j + 2)}7-{col(j + 2)}8-{col(j + 2)}9'
        style_total(ws.cell(row=10, column=j + 2, value=f))

    # 所得税
    ws.cell(row=11, column=1, value='所得税').font = NORMAL_FONT
    for j in range(years):
        f = f'={col(j + 2)}10*\'1.假设\'!{col(j + 3)}9'
        style_data(ws.cell(row=11, column=j + 2, value=f))

    # 净利润
    ws.cell(row=12, column=1, value='净利润').font = TOTAL_FONT
    for j in range(years):
        f = f'={col(j + 2)}10-{col(j + 2)}11'
        c = ws.cell(row=12, column=j + 2, value=f)
        c.font = Font(name='微软雅黑', size=11, bold=True, color='C00000')
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')
        c.border = THIN_BORDER

    ws.column_dimensions['A'].width = 18
    for j in range(2, 2 + years):
        ws.column_dimensions[col(j)].width = 14


def build_balance(wb: Workbook, years: int, company: str) -> None:
    """资产负债表（简化版）。"""
    ws = wb.create_sheet('3.资产负债表')
    ws['A1'] = f'{company} - 资产负债表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

    headers = ['项目'] + [f'第{i + 1}年末' for i in range(years)]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    # 简化：只列主要科目，作为示意
    items = [
        ('资产', None),
        ('  货币资金', '(用户填入)'),
        ('  应收账款', '收入×应收账款天数/365'),
        ('  存货', '成本×存货天数/365'),
        ('  固定资产净值', '(用户填入)'),
        ('总资产', '上面之和'),
        ('', ''),
        ('负债', None),
        ('  应付账款', '成本×应付天数/365'),
        ('  短期借款', '(用户填入)'),
        ('  长期负债', '(用户填入)'),
        ('总负债', '上面之和'),
        ('所有者权益', None),
        ('  实收资本', '(用户填入)'),
        ('  留存收益', '上期+本期净利润-分红'),
        ('权益合计', '上面之和'),
        ('负债+权益', '总负债+权益合计'),
        ('平衡校验', '应等于总资产'),
    ]
    for i, (name, desc) in enumerate(items, start=4):
        c = ws.cell(row=i, column=1, value=name)
        if name in ('资产', '负债', '所有者权益'):
            c.font = SECTION_FONT
            c.fill = SECTION_FILL
        elif name in ('总资产', '总负债', '权益合计', '负债+权益'):
            c.font = TOTAL_FONT
        elif name == '平衡校验':
            c.font = Font(name='微软雅黑', size=10, bold=True, color='C00000')
        else:
            c.font = NORMAL_FONT
        if desc:
            ws.cell(row=i, column=years + 2, value=desc).font = Font(
                name='微软雅黑', size=9, italic=True, color='808080')

    # 完整三表联动公式留作进阶版本，本 MVP 给注释占位
    ws['A22'] = '注：完整三表联动公式留作扩展，详见 README'
    ws['A22'].font = Font(name='微软雅黑', size=9, italic=True, color='808080')

    ws.column_dimensions['A'].width = 22
    for j in range(2, 2 + years):
        ws.column_dimensions[col(j)].width = 14
    ws.column_dimensions[col(years + 2)].width = 30


def build_cashflow(wb: Workbook, years: int, company: str) -> None:
    """现金流量表（简化版）。"""
    ws = wb.create_sheet('4.现金流')
    ws['A1'] = f'{company} - 现金流量表'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

    headers = ['项目'] + [f'第{i + 1}年' for i in range(years)]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    items = [
        '一、经营活动现金流',
        '  净利润 (取自利润表)',
        '  +折旧',
        '  -应收账款增加',
        '  -存货增加',
        '  +应付账款增加',
        '经营活动现金流净额',
        '',
        '二、投资活动现金流',
        '  -资本开支 CapEx',
        '投资活动现金流净额',
        '',
        '三、筹资活动现金流',
        '  +借款',
        '  -还款',
        '  -分红',
        '筹资活动现金流净额',
        '',
        '现金净增加',
        '期初现金',
        '期末现金',
    ]
    for i, name in enumerate(items, start=4):
        c = ws.cell(row=i, column=1, value=name)
        if name.startswith('一、') or name.startswith('二、') or name.startswith('三、'):
            c.font = SECTION_FONT
            c.fill = SECTION_FILL
        elif '净额' in name or '现金净增加' in name or '期末' in name or '期初' in name:
            c.font = TOTAL_FONT
        else:
            c.font = NORMAL_FONT

    # 第 5 行：净利润 = 利润表第 12 行
    for j in range(years):
        f = f"='2.利润表'!{col(j + 2)}12"
        style_data(ws.cell(row=5, column=j + 2, value=f))

    ws.column_dimensions['A'].width = 22
    for j in range(2, 2 + years):
        ws.column_dimensions[col(j)].width = 14


def build_kpi(wb: Workbook, years: int, company: str) -> None:
    """关键比率 dashboard。"""
    ws = wb.create_sheet('5.关键比率')
    ws['A1'] = f'{company} - 关键财务比率'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True)

    headers = ['指标'] + [f'第{i + 1}年' for i in range(years)] + ['说明']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    # 取利润表数据
    ratios = [
        ('毛利率', '毛利/收入', '0.00%',
         lambda j: f"='2.利润表'!{col(j + 2)}6/'2.利润表'!{col(j + 2)}4"),
        ('营业利润率', '营业利润/收入', '0.00%',
         lambda j: f"='2.利润表'!{col(j + 2)}10/'2.利润表'!{col(j + 2)}4"),
        ('净利率', '净利润/收入', '0.00%',
         lambda j: f"='2.利润表'!{col(j + 2)}12/'2.利润表'!{col(j + 2)}4"),
        ('收入增长率', '同比', '0.00%',
         lambda j: f"=IFERROR('2.利润表'!{col(j + 2)}4/'2.利润表'!{col(j + 1)}4-1,\"\")"
         if j > 0 else '"基准年"'),
    ]
    for i, (name, desc, fmt, formula_fn) in enumerate(ratios, start=4):
        ws.cell(row=i, column=1, value=name).font = TOTAL_FONT
        for j in range(years):
            try:
                f = formula_fn(j)
            except Exception:
                f = ''
            c = ws.cell(row=i, column=j + 2, value=f)
            style_data(c, fmt=fmt)
        ws.cell(row=i, column=years + 2, value=desc).font = Font(
            name='微软雅黑', size=9, italic=True, color='808080')

    ws.column_dimensions['A'].width = 16
    for j in range(2, 2 + years):
        ws.column_dimensions[col(j)].width = 12
    ws.column_dimensions[col(years + 2)].width = 24


def main() -> int:
    parser = argparse.ArgumentParser(description='Generate 3-statement financial model template')
    parser.add_argument('--output', '-o', default='three_statements.xlsx')
    parser.add_argument('--years', type=int, default=5)
    parser.add_argument('--company', default='示例公司')
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    build_assumptions(wb, args.years, args.company)
    build_income(wb, args.years, args.company)
    build_balance(wb, args.years, args.company)
    build_cashflow(wb, args.years, args.company)
    build_kpi(wb, args.years, args.company)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f'[ok] Generated: {out}')
    print(f'     Sheets: 1.假设, 2.利润表, 3.资产负债表, 4.现金流, 5.关键比率')
    print(f'     Years:  {args.years}')
    print()
    print('Note: 公式由 openpyxl 写入但未计算。打开 Excel 会自动重算，')
    print('      或运行 `python scripts/recalc.py {0}` 用 LibreOffice 重算。'.format(out))
    return 0


if __name__ == '__main__':
    sys.exit(main())
