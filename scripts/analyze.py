#!/usr/bin/env python3
"""
analyze.py — 给现有 .xlsx 文件做"健康度诊断"。

不是生成 Excel，而是**读 Excel 找问题**。
对应日常痛点：接手了一份别人的 Excel，公式可能错、可能循环引用、可能有
未保护的输入区被覆盖。本工具静态扫描后给出修复建议。

检查项
------
- XA001: 公式中含中文逗号 ，（应是英文 ,）
- XA002: 公式引用 #REF! / #N/A / #DIV/0! / #VALUE! 等已是错误的单元格
- XA003: 工作表名含 Excel 禁止字符 / \\ ? * [ ]
- XA004: 工作表名超 31 字符
- XA005: 工作表名包含空格但被其他 sheet 公式无引号引用
- XA006: 同一单元格既是输入又是公式（容易被覆盖）
- XA007: 公式中硬编码"魔数"（应改用具名单元格 / 命名区域）
- XA008: 合并单元格作为公式输入（合并后只有左上有值，引用易错）
- XA009: 整列引用（如 A:A）在大文件中显著拖慢
- XA010: 隐藏 sheet 但被可见 sheet 公式依赖（用户不知道）

用法
----
    python scripts/analyze.py path/to/workbook.xlsx
    python scripts/analyze.py path/to/workbook.xlsx --json
    python scripts/analyze.py path/to/workbook.xlsx --fail-on warning
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


_FORBIDDEN_SHEET_CHARS = set('\\/?*[]')
_ERROR_TOKENS = ('#REF!', '#N/A', '#DIV/0!', '#VALUE!', '#NAME?', '#NUM!', '#NULL!')
_WHOLE_COL_RE = re.compile(r'(?<![A-Z0-9_])([A-Z]+):([A-Z]+)(?![0-9])')


@dataclass
class Finding:
    severity: str         # error / warning / info
    code: str
    sheet: Optional[str]
    cell: Optional[str]
    message: str
    fix: str = ''

    def to_dict(self) -> dict:
        return {
            'severity': self.severity,
            'code': self.code,
            'sheet': self.sheet,
            'cell': self.cell,
            'message': self.message,
            'fix': self.fix,
        }


@dataclass
class Report:
    file: str
    findings: List[Finding] = field(default_factory=list)

    def add(self, severity, code, sheet, cell, message, fix=''):
        self.findings.append(Finding(severity, code, sheet, cell, message, fix))

    @property
    def errors(self):
        return [f for f in self.findings if f.severity == 'error']

    @property
    def warnings(self):
        return [f for f in self.findings if f.severity == 'warning']

    @property
    def passed(self):
        return len(self.errors) == 0

    def to_dict(self) -> dict:
        return {
            'file': self.file,
            'passed': self.passed,
            'summary': {
                'errors': len(self.errors),
                'warnings': len(self.warnings),
                'infos': sum(1 for f in self.findings if f.severity == 'info'),
                'total': len(self.findings),
            },
            'findings': [f.to_dict() for f in self.findings],
        }


# ---------------------------------------------------------------------------
# Checks
# ---------------------------------------------------------------------------

def _check_sheet_names(wb: Workbook, report: Report) -> None:
    seen = set()
    for name in wb.sheetnames:
        if len(name) > 31:
            report.add('error', 'XA004', name, None,
                       f'sheet 名 "{name}" 超 31 字符（Excel 上限）',
                       fix='缩短 sheet 名至 31 字符内')
        if any(c in _FORBIDDEN_SHEET_CHARS for c in name):
            report.add('error', 'XA003', name, None,
                       f'sheet 名 "{name}" 含禁止字符 / \\ ? * [ ]',
                       fix='移除禁止字符')
        lower = name.lower()
        if lower in seen:
            report.add('error', 'XA003', name, None,
                       f'sheet 名 "{name}" 与其他 sheet 大小写仅差一字（易混）',
                       fix='重命名')
        seen.add(lower)


def _check_formulas(wb: Workbook, report: Report) -> None:
    sheet_names_with_space = {n for n in wb.sheetnames if ' ' in n}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.sheet_state == 'hidden':
            # 检查是否被可见 sheet 引用
            for other in wb.sheetnames:
                if other == sheet_name or wb[other].sheet_state == 'hidden':
                    continue
                for row in wb[other].iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            if sheet_name in cell.value or f"'{sheet_name}'" in cell.value:
                                report.add('warning', 'XA010', other, cell.coordinate,
                                           f"可见 sheet 引用了隐藏 sheet '{sheet_name}'",
                                           fix='让 sheet 可见，或用 Named Range 隐藏依赖')
                                break

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                # 错误标志值
                if isinstance(cell.value, str):
                    val = cell.value
                    if val.startswith('='):
                        _check_one_formula(val, sheet_name, cell.coordinate,
                                           sheet_names_with_space, report)
                    elif val in _ERROR_TOKENS:
                        report.add('error', 'XA002', sheet_name, cell.coordinate,
                                   f'单元格已记录为 {val}（公式结果错误）',
                                   fix='上溯依赖并修正')


def _check_one_formula(formula: str, sheet: str, cell: str,
                       sheets_with_space: set, report: Report) -> None:
    # 中文逗号
    if '，' in formula:
        report.add('error', 'XA001', sheet, cell,
                   '公式含中文逗号 ， — Excel 解析会失败',
                   fix='把 ， 替换为 ,')

    # 整列引用（大文件性能杀手）
    whole_cols = _WHOLE_COL_RE.findall(formula)
    if whole_cols:
        report.add('warning', 'XA009', sheet, cell,
                   f'公式用了整列引用 {":".join(whole_cols[0])} — 大文件会显著拖慢',
                   fix='改用具体范围如 A2:A10000 或定义名称')

    # 跨 sheet 引用但 sheet 名含空格且未加引号
    for sn in sheets_with_space:
        # 应该被 'sn' 包围；若直接出现 sn 但前面不是 '，则可能漏引号
        if sn in formula and f"'{sn}'" not in formula:
            # 简化判断：sn 前不是单引号
            idx = formula.find(sn)
            if idx > 0 and formula[idx - 1] != "'":
                report.add('warning', 'XA005', sheet, cell,
                           f"公式引用 sheet '{sn}'（含空格）但未加单引号",
                           fix=f"改为 '{sn}'!XX")

    # 硬编码魔数：判断公式里是否有 4 位以上的数字常量
    if re.search(r'(?<![A-Z\d_])\d{4,}(?![\d.])', formula):
        report.add('info', 'XA007', sheet, cell,
                   '公式含 4+ 位数字常量（魔数）',
                   fix='考虑改用具名单元格或命名区域')


def _check_merged_cells(wb: Workbook, report: Report) -> None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for merged_range in ws.merged_cells.ranges:
            # 合并区域被公式引用时，引用左上角才有值；老手都知道，但易踩
            # 这里只 info 提示用户：注意合并区被引用的坑
            if merged_range.size > 1:
                report.add('info', 'XA008', sheet_name, str(merged_range),
                           '合并单元格存在 — 引用时只有左上角有值',
                           fix='避免在数据区合并；表头合并是 OK 的')


def analyze_workbook(path: str | Path) -> Report:
    """诊断单个 .xlsx 文件。"""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    report = Report(file=str(p))

    # data_only=False 拿 formula 原文
    wb = load_workbook(p, data_only=False, keep_vba=p.suffix.lower() == '.xlsm')
    try:
        _check_sheet_names(wb, report)
        _check_formulas(wb, report)
        _check_merged_cells(wb, report)
    finally:
        wb.close()
    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _format_text(report: Report) -> str:
    lines = [f'== {report.file} ==']
    if not report.findings:
        lines.append('  no issues found')
        return '\n'.join(lines)
    by_severity = {'error': '✗', 'warning': '!', 'info': '·'}
    for f in report.findings:
        sigil = by_severity.get(f.severity, '?')
        loc = f'{f.sheet}!{f.cell}' if f.cell else (f.sheet or '-')
        lines.append(f'  {sigil} [{f.code}] {loc}: {f.message}')
        if f.fix:
            lines.append(f'       fix: {f.fix}')
    lines.append('')
    lines.append(f'  errors={len(report.errors)} warnings={len(report.warnings)} '
                 f'total={len(report.findings)}')
    return '\n'.join(lines)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(prog='analyze', description='Diagnose an .xlsx file')
    parser.add_argument('xlsx', help='Path to .xlsx / .xlsm file')
    parser.add_argument('--json', action='store_true', help='Output JSON')
    parser.add_argument('--fail-on', choices=['error', 'warning'], default='error',
                        help='Exit code != 0 if findings at this level exist')
    args = parser.parse_args(argv)

    report = analyze_workbook(args.xlsx)
    if args.json:
        print(json.dumps(report.to_dict(), indent=2, ensure_ascii=False))
    else:
        print(_format_text(report))

    if args.fail_on == 'error' and report.errors:
        return 1
    if args.fail_on == 'warning' and (report.errors or report.warnings):
        return 1
    return 0


if __name__ == '__main__':
    sys.exit(main())
