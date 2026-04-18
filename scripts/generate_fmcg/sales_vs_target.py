#!/usr/bin/env python3
"""
sales_vs_target.py — 快消行业销售达成 vs 目标 Excel 模板

生成包含以下 sheet 的模板：
- Sheet 1: 月度目标设定（按 SKU + 渠道）
- Sheet 2: 实际销售数据（每日录入）
- Sheet 3: 达成率 Dashboard（按 SKU / 按渠道 / 按月汇总）
- Sheet 4: SKU ABC 分类（Pareto 80/20 分析）

用法:
    python scripts/generate_fmcg/sales_vs_target.py --output sales.xlsx
    python scripts/generate_fmcg/sales_vs_target.py --skus 30 --channels 5
"""
from __future__ import annotations

import argparse
import random
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')  # 快消蓝
NORMAL_FONT = Font(name='微软雅黑', size=10)
TOTAL_FONT = Font(name='微软雅黑', size=10, bold=True)
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')


def col(n: int) -> str:
    return get_column_letter(n)


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')


def build_targets(wb: Workbook, n_skus: int, channels: list[str]) -> None:
    ws = wb.create_sheet('1.月度目标')
    ws['A1'] = '月度销售目标设定（按 SKU × 渠道）'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')

    headers = ['SKU 代码', 'SKU 名称', '类目'] + channels + ['月度合计']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    categories = ['食品', '日化', '饮料', '清洁', '美妆']
    random.seed(42)

    for i in range(n_skus):
        row = i + 4
        ws.cell(row=row, column=1, value=f'SKU{i + 1:04d}').font = NORMAL_FONT
        ws.cell(row=row, column=2, value=f'产品 {i + 1}').font = NORMAL_FONT
        ws.cell(row=row, column=3, value=random.choice(categories)).font = NORMAL_FONT
        for j, _ch in enumerate(channels, start=4):
            target = random.randint(1000, 50000)
            c = ws.cell(row=row, column=j, value=target)
            c.fill = INPUT_FILL
            c.font = NORMAL_FONT
            c.number_format = '#,##0'
            c.alignment = Alignment(horizontal='right')
        # 合计公式
        last_ch_col = col(3 + len(channels))
        f = f'=SUM(D{row}:{last_ch_col}{row})'
        c = ws.cell(row=row, column=4 + len(channels), value=f)
        c.font = TOTAL_FONT
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')

    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    for j in range(4, 4 + len(channels) + 1):
        ws.column_dimensions[col(j)].width = 12


def build_actuals(wb: Workbook, n_skus: int, channels: list[str], days: int = 30) -> None:
    ws = wb.create_sheet('2.实际销售')
    ws['A1'] = f'实际销售数据（最近 {days} 天）'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')

    headers = ['日期', 'SKU 代码', '渠道', '销量', '单价', '销售额']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    today = date.today()
    random.seed(42)

    row = 4
    for d_back in range(days):
        d = today - timedelta(days=d_back)
        # 每天约 30 笔数据
        for _ in range(min(30, n_skus * 2)):
            sku_idx = random.randint(1, n_skus)
            ch = random.choice(channels)
            qty = random.randint(10, 500)
            price = random.uniform(5, 100)

            ws.cell(row=row, column=1, value=d).number_format = 'yyyy-mm-dd'
            ws.cell(row=row, column=2, value=f'SKU{sku_idx:04d}')
            ws.cell(row=row, column=3, value=ch)
            ws.cell(row=row, column=4, value=qty)
            c = ws.cell(row=row, column=5, value=round(price, 2))
            c.number_format = '#,##0.00'
            # 销售额 = 销量 * 单价
            ws.cell(row=row, column=6, value=f'=D{row}*E{row}').number_format = '#,##0.00'
            row += 1

    for j in range(1, 7):
        ws.column_dimensions[col(j)].width = 14


def build_dashboard(wb: Workbook, channels: list[str]) -> None:
    ws = wb.create_sheet('3.达成率Dashboard')
    ws['A1'] = '销售达成率 Dashboard'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')

    ws['A3'] = '本月汇总'
    ws['A3'].font = TOTAL_FONT

    headers = ['渠道', '目标', '实际', '达成率', '差距']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=j, value=h))

    for i, ch in enumerate(channels, start=5):
        ws.cell(row=i, column=1, value=ch).font = NORMAL_FONT
        # 目标：从 sheet 1 用 SUMIFS 取该渠道总目标
        # 简化：目标列在 sheet 1 的位置取决于 channels 位置
        ch_col = col(4 + channels.index(ch))
        f_target = f"=SUM('1.月度目标'!{ch_col}4:{ch_col}1000)"
        c = ws.cell(row=i, column=2, value=f_target)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')

        # 实际：从 sheet 2 SUMIFS 该渠道销售额
        f_actual = f"=SUMIF('2.实际销售'!C:C,A{i},'2.实际销售'!F:F)"
        c = ws.cell(row=i, column=3, value=f_actual)
        c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')

        # 达成率
        c = ws.cell(row=i, column=4, value=f'=IFERROR(C{i}/B{i},"")')
        c.number_format = '0.0%'
        c.alignment = Alignment(horizontal='right')

        # 差距
        c = ws.cell(row=i, column=5, value=f'=C{i}-B{i}')
        c.number_format = '#,##0;[红色]-#,##0'
        c.alignment = Alignment(horizontal='right')

    # 条件格式：达成率 < 80% 红，>= 100% 绿
    last_row = 4 + len(channels)
    ws.conditional_formatting.add(
        f'D5:D{last_row}',
        CellIsRule(operator='lessThan', formula=['0.8'],
                   fill=PatternFill('solid', fgColor='FFC7CE'),
                   font=Font(color='9C0006'))
    )
    ws.conditional_formatting.add(
        f'D5:D{last_row}',
        CellIsRule(operator='greaterThanOrEqual', formula=['1.0'],
                   fill=PatternFill('solid', fgColor='C6EFCE'),
                   font=Font(color='006100'))
    )

    for j in range(1, 6):
        ws.column_dimensions[col(j)].width = 14


def build_abc(wb: Workbook, n_skus: int) -> None:
    ws = wb.create_sheet('4.SKU_ABC分类')
    ws['A1'] = 'SKU ABC 分类（Pareto 80/20）'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='2E75B6')

    note = (
        '操作说明：\n'
        '1. 在「销售额」列填入或公式取自 sheet 2 的 SKU 汇总\n'
        '2. 按销售额降序排列\n'
        '3. 累计占比 ≤ 80% → A 类；80%-95% → B 类；> 95% → C 类'
    )
    ws['A2'] = note
    ws['A2'].font = Font(name='微软雅黑', size=9, italic=True, color='808080')
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:E4')

    headers = ['SKU 代码', '销售额', '占比', '累计占比', 'ABC 分类']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=6, column=j, value=h))

    # 留白让用户填，给前 5 行公式示例
    for i in range(min(n_skus, 5)):
        row = 7 + i
        ws.cell(row=row, column=1, value=f'SKU{i + 1:04d}').font = NORMAL_FONT
        # 销售额：从 sheet 2 用 SUMIF
        ws.cell(row=row, column=2,
                value=f"=SUMIF('2.实际销售'!B:B,A{row},'2.实际销售'!F:F)"
                ).number_format = '#,##0'
        # 占比 = 单 SKU / 总和
        ws.cell(row=row, column=3,
                value=f'=IFERROR(B{row}/SUM(B$7:B$1000),"")'
                ).number_format = '0.0%'
        # 累计占比
        ws.cell(row=row, column=4,
                value=f'=SUM(C$7:C{row})'
                ).number_format = '0.0%'
        # ABC 分类
        ws.cell(row=row, column=5,
                value=f'=IF(D{row}<=0.8,"A",IF(D{row}<=0.95,"B","C"))'
                )

    for j in range(1, 6):
        ws.column_dimensions[col(j)].width = 14


def main() -> int:
    parser = argparse.ArgumentParser(description='Generate FMCG sales vs target template')
    parser.add_argument('--output', '-o', default='sales_vs_target.xlsx')
    parser.add_argument('--skus', type=int, default=20, help='Number of SKUs (default 20)')
    parser.add_argument('--channels', type=int, default=4, help='Number of channels (default 4)')
    args = parser.parse_args()

    channel_names = ['现代渠道', '传统渠道', '电商', 'KA', '便利店', '专业渠道'][:args.channels]
    if len(channel_names) < args.channels:
        channel_names += [f'渠道{i + 1}' for i in range(len(channel_names), args.channels)]

    wb = Workbook()
    wb.remove(wb.active)

    build_targets(wb, args.skus, channel_names)
    build_actuals(wb, args.skus, channel_names, days=30)
    build_dashboard(wb, channel_names)
    build_abc(wb, args.skus)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f'[ok] Generated: {out}')
    print(f'     Sheets: 1.月度目标, 2.实际销售, 3.达成率Dashboard, 4.SKU_ABC分类')
    print(f'     Channels: {", ".join(channel_names)}')
    print(f'     SKUs: {args.skus}')
    print()
    print('Note: 公式由 openpyxl 写入但未计算。打开 Excel 会自动重算。')
    return 0


if __name__ == '__main__':
    sys.exit(main())
