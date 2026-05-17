#!/usr/bin/env python3
"""
unit_economics.py — SaaS 单位经济模型生成器（v2 新增 vertical）

生成一份 SaaS 公司经典经济模型 Excel：

- Sheet 1: Assumptions（假设输入）— ARPU / Churn / CAC / Gross Margin
- Sheet 2: Cohort Retention — 月度 cohort 留存矩阵 + 累计收入
- Sheet 3: Unit Economics — LTV / CAC payback / LTV-to-CAC ratio
- Sheet 4: P&L Projection — 3 年收入 / 毛利 / S&M 投入 / 净亏损 / 现金需求
- Sheet 5: KPI Dashboard — 关键指标可视化（用 Excel 内置 chart）

公式联动：
- Churn → cohort retention curve
- ARPU × retained customers → 月度 MRR
- LTV = ARPU × Gross Margin / Monthly Churn
- LTV/CAC ratio 用于评估单位经济是否健康（行业经验 > 3 才算好）

用法：
    python scripts/generate_saas/unit_economics.py --output saas_model.xlsx
    python scripts/generate_saas/unit_economics.py --output x.xlsx --months 36 \\
        --arpu 100 --gross-margin 0.8 --monthly-churn 0.03 --cac 1500
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2E75B6')
SECTION_FONT = Font(name='微软雅黑', size=10, bold=True, color='1F4E78')
SECTION_FILL = PatternFill('solid', fgColor='DDEBF7')
TOTAL_FONT = Font(name='微软雅黑', size=10, bold=True)
TOTAL_FILL = PatternFill('solid', fgColor='FCE4D6')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')
THIN = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def _section(ws, row: int, title: str, span: int = 12) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    ws.merge_cells(start_row=row, end_row=row, start_column=1, end_column=span)


def _build_assumptions(wb: Workbook, args) -> None:
    ws = wb.create_sheet('Assumptions')
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 16

    inputs = [
        ('ARPU 每月每客户（元）', args.arpu, 'B2'),
        ('Gross Margin（毛利率）', args.gross_margin, 'B3'),
        ('Monthly Churn（月流失率）', args.monthly_churn, 'B4'),
        ('CAC 每客户获取成本（元）', args.cac, 'B5'),
        ('Initial New Customers / Month', args.new_customers, 'B6'),
        ('Customer Growth Rate / Month', args.growth_rate, 'B7'),
        ('Months 模拟周期', args.months, 'B8'),
        ('Discount Rate（年化）', args.discount_rate, 'B9'),
    ]
    ws['A1'] = '输入假设（黄色为可调）'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:B1')

    for i, (label, value, _) in enumerate(inputs, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(name='微软雅黑', size=10)
        c = ws.cell(row=i, column=2, value=value)
        c.fill = INPUT_FILL
        c.font = Font(name='微软雅黑', size=10, bold=True)
        c.border = THIN
        if 'Rate' in label or 'Churn' in label or 'Margin' in label:
            c.number_format = '0.00%'
        elif 'Month' in label and 'ARPU' not in label and '模拟' not in label:
            c.number_format = '#,##0'
        elif '模拟' in label:
            c.number_format = '0'
        else:
            c.number_format = '#,##0.00'

    # 衍生指标
    ws['A11'] = '衍生指标（自动计算）'
    ws['A11'].font = SECTION_FONT
    ws['A11'].fill = SECTION_FILL
    ws.merge_cells('A11:B11')

    derived = [
        ('LTV = ARPU × GM ÷ Churn', '=B2*B3/B4'),
        ('LTV/CAC Ratio', '=B12/B5'),
        ('CAC Payback (月)', '=B5/(B2*B3)'),
        ('Annual Churn (近似)', '=1-(1-B4)^12'),
    ]
    for i, (label, formula) in enumerate(derived, start=12):
        ws.cell(row=i, column=1, value=label).font = Font(name='微软雅黑', size=10)
        c = ws.cell(row=i, column=2, value=formula)
        c.font = TOTAL_FONT
        c.fill = TOTAL_FILL
        c.border = THIN
        if 'Ratio' in label:
            c.number_format = '0.00"x"'
        elif 'Payback' in label:
            c.number_format = '0.0"月"'
        elif 'Churn' in label:
            c.number_format = '0.0%'
        else:
            c.number_format = '#,##0.00'


def _build_cohort(wb: Workbook, args) -> None:
    ws = wb.create_sheet('Cohort')
    ws['A1'] = 'Cohort 月度留存矩阵（行=新客户起始月，列=月偏移）'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(args.months + 2, 14))

    # 表头
    ws.cell(row=2, column=1, value='Cohort 起始月').font = TOTAL_FONT
    ws.cell(row=2, column=2, value='cohort 大小').font = TOTAL_FONT
    max_offset = min(args.months, 12)  # 展示 12 列，超出截断
    for offset in range(max_offset + 1):
        ws.cell(row=2, column=3 + offset, value=f'M+{offset}').font = TOTAL_FONT

    # 数据
    for m in range(args.months):
        row = 3 + m
        ws.cell(row=row, column=1, value=f'Month {m + 1}').font = Font(name='微软雅黑', size=10)
        # cohort 大小公式：初始 × (1 + 增长率)^m
        ws.cell(row=row, column=2, value=f'=Assumptions!$B$6*(1+Assumptions!$B$7)^{m}').number_format = '#,##0'
        for offset in range(max_offset + 1):
            # cohort × (1 - churn)^offset
            ws.cell(row=row, column=3 + offset,
                    value=f'=$B{row}*(1-Assumptions!$B$4)^{offset}').number_format = '#,##0'


def _build_unit_economics(wb: Workbook, args) -> None:
    ws = wb.create_sheet('UnitEconomics')
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16

    rows = [
        ('CAC（每客户获取成本）', '=Assumptions!B5'),
        ('ARPU（每月）', '=Assumptions!B2'),
        ('Gross Margin', '=Assumptions!B3'),
        ('Monthly Churn', '=Assumptions!B4'),
        ('Contribution Margin per Month', '=B3*B4'),  # ARPU * GM ... no wait
    ]
    # 重新定义：让公式严格清楚
    rows = [
        ('CAC', '=Assumptions!B5'),
        ('ARPU/月', '=Assumptions!B2'),
        ('Gross Margin', '=Assumptions!B3'),
        ('Monthly Churn', '=Assumptions!B4'),
        ('月度贡献毛利 = ARPU × GM', '=B3*B4'),
        ('LTV = 月毛利 / Churn', '=B6/B5'),
        ('LTV / CAC', '=B7/B2'),
        ('CAC Payback（月）', '=B2/B6'),
        ('Magic Number = NewMRR/CAC', '=Assumptions!B2*Assumptions!B6/Assumptions!B5'),
    ]
    ws['A1'] = 'Unit Economics（关键单位经济指标）'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:B1')

    for i, (label, formula) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label).font = Font(name='微软雅黑', size=10)
        c = ws.cell(row=i, column=2, value=formula)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = THIN
        if 'Margin' in label or 'Churn' in label:
            c.number_format = '0.0%'
        elif 'LTV / CAC' in label or 'Magic' in label:
            c.number_format = '0.00"x"'
        elif 'Payback' in label:
            c.number_format = '0.0"月"'
        else:
            c.number_format = '#,##0.00'

    # 健康度提示
    ws['A12'] = '健康度（基于业内 benchmark）'
    ws['A12'].font = SECTION_FONT
    ws['A12'].fill = SECTION_FILL
    ws.merge_cells('A12:B12')
    ws['A13'] = 'LTV/CAC >3 才视为健康'
    ws['B13'] = '=IF(B8>=3, "✓ Healthy", IF(B8>=1, "! Watch", "✗ Burning"))'
    ws['A14'] = 'Payback <12 月较好'
    ws['B14'] = '=IF(B9<=12, "✓ Fast", IF(B9<=24, "! OK", "✗ Slow"))'


def _build_pl(wb: Workbook, args) -> None:
    ws = wb.create_sheet('PL_Projection')
    ws.column_dimensions['A'].width = 26
    max_show = min(args.months, 36)

    # 表头
    ws.cell(row=1, column=1, value='科目').font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    for m in range(max_show):
        c = ws.cell(row=1, column=2 + m, value=f'M{m + 1}')
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    # 行：active customers / MRR / 毛利 / S&M cost / 净利
    rows_def = [
        ('Active Customers', None),
        ('MRR', None),
        ('Gross Profit', None),
        ('S&M Cost (CAC × New)', None),
        ('Net Margin', None),
    ]
    for i, (label, _) in enumerate(rows_def, start=2):
        ws.cell(row=i, column=1, value=label).font = TOTAL_FONT

    for m in range(max_show):
        col_letter = get_column_letter(2 + m)
        # Active = sum of cohorts retained，简化用 Cohort 表的对角线和
        # = SUMPRODUCT(cohort 大小 × (1-churn)^(m-startMonth))
        # 这里简化为：取 cohort 表第 (m+1) 列的总和
        if m == 0:
            ws.cell(row=2, column=2 + m,
                    value='=Assumptions!B6')   # 第 1 月新客户
        else:
            ws.cell(row=2, column=2 + m,
                    value=f'={col_letter}2-{get_column_letter(2 + m - 1)}2*Assumptions!B4+Assumptions!B6*(1+Assumptions!B7)^{m}')
        ws.cell(row=2, column=2 + m).number_format = '#,##0'

        ws.cell(row=3, column=2 + m, value=f'={col_letter}2*Assumptions!$B$2').number_format = '#,##0'
        ws.cell(row=4, column=2 + m, value=f'={col_letter}3*Assumptions!$B$3').number_format = '#,##0'
        ws.cell(row=5, column=2 + m,
                value=f'=Assumptions!$B$6*(1+Assumptions!$B$7)^{m}*Assumptions!$B$5').number_format = '#,##0'
        ws.cell(row=6, column=2 + m, value=f'={col_letter}4-{col_letter}5').number_format = '#,##0'


def _build_kpi_dashboard(wb: Workbook, args) -> None:
    ws = wb.create_sheet('Dashboard')
    ws['A1'] = 'KPI Dashboard'
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:F1')

    metrics = [
        ('LTV', '=UnitEconomics!B7'),
        ('CAC', '=UnitEconomics!B2'),
        ('LTV/CAC', '=UnitEconomics!B8'),
        ('Payback (月)', '=UnitEconomics!B9'),
        ('Annual Churn', '=Assumptions!B15'),
        ('Magic Number', '=UnitEconomics!B10'),
    ]
    for i, (k, v) in enumerate(metrics, start=2):
        ws.cell(row=i, column=1, value=k).font = TOTAL_FONT
        c = ws.cell(row=i, column=2, value=v)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT


def generate(args: argparse.Namespace) -> Path:
    wb = Workbook()
    # 删除默认 sheet
    default = wb.active
    wb.remove(default)

    _build_assumptions(wb, args)
    _build_cohort(wb, args)
    _build_unit_economics(wb, args)
    _build_pl(wb, args)
    _build_kpi_dashboard(wb, args)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(prog='unit_economics',
                                     description='SaaS 单位经济模型生成器')
    parser.add_argument('--output', '-o', default='saas_unit_economics.xlsx')
    parser.add_argument('--months', type=int, default=24)
    parser.add_argument('--arpu', type=float, default=100.0, help='ARPU 每月每客户')
    parser.add_argument('--gross-margin', type=float, default=0.80)
    parser.add_argument('--monthly-churn', type=float, default=0.03)
    parser.add_argument('--cac', type=float, default=1500.0)
    parser.add_argument('--new-customers', type=int, default=100,
                        help='初始每月新客户数')
    parser.add_argument('--growth-rate', type=float, default=0.10,
                        help='新客户每月增长率')
    parser.add_argument('--discount-rate', type=float, default=0.10,
                        help='年化折现率（用于 NPV，可暂不使用）')
    args = parser.parse_args(argv)

    path = generate(args)
    # 用 stdout.buffer 写 UTF-8，避免 Windows GBK 控制台对 ✓ 的 UnicodeEncodeError
    msg = f'[OK] Generated: {path}'
    try:
        print(msg)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((msg + '\n').encode('utf-8'))
    return 0


if __name__ == '__main__':
    sys.exit(main())
