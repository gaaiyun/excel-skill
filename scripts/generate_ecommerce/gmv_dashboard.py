#!/usr/bin/env python3
"""
gmv_dashboard.py — 电商 GMV Dashboard 模板生成器

生成一个完整的电商 GMV 分析模板，包含：
- Sheet 1: 数据录入（订单明细，30 天 sample）
- Sheet 2: KPI 看板（GMV / 订单数 / 客单价 / 转化率，按日 + 移动平均）
- Sheet 3: 类目分析（销售排名 + ABC 分类）
- Sheet 4: 漏斗（曝光 → 点击 → 加购 → 下单 → 支付）

用法:
    python scripts/generate_ecommerce/gmv_dashboard.py --output gmv.xlsx
    python scripts/generate_ecommerce/gmv_dashboard.py --output gmv.xlsx --days 90
"""
from __future__ import annotations

import argparse
import random
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


# ============ 样式 ============

HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
NORMAL_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER


def style_data(cell, number_format=None):
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = THIN_BORDER
    if number_format:
        cell.number_format = number_format


# ============ Sheet 1: 数据录入 ============

CATEGORIES = ['服装', '美妆', '数码', '家居', '食品', '母婴', '运动']

def make_orders(num_days: int) -> list[dict]:
    """生成 num_days 天的样本订单数据。"""
    random.seed(42)  # 可复现
    orders = []
    today = date.today()
    for d in range(num_days):
        day = today - timedelta(days=num_days - d - 1)
        # 每天 50-150 单
        n_orders = random.randint(50, 150)
        for _ in range(n_orders):
            cat = random.choice(CATEGORIES)
            price = round(random.uniform(20, 800), 2)
            qty = random.randint(1, 5)
            orders.append({
                'date': day,
                'order_id': f'ORD{day.strftime("%Y%m%d")}{random.randint(1000, 9999)}',
                'category': cat,
                'price': price,
                'qty': qty,
                'amount': round(price * qty, 2),
            })
    return orders


def build_data_sheet(wb: Workbook, orders: list[dict]) -> None:
    ws = wb.create_sheet('数据录入', 0)
    headers = ['日期', '订单号', '商品类目', '单价(元)', '数量', '订单金额(元)']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)

    for row_idx, order in enumerate(orders, start=2):
        ws.cell(row=row_idx, column=1, value=order['date'])
        style_data(ws.cell(row=row_idx, column=1), 'yyyy-mm-dd')
        ws.cell(row=row_idx, column=2, value=order['order_id']); style_data(ws.cell(row=row_idx, column=2))
        ws.cell(row=row_idx, column=3, value=order['category']); style_data(ws.cell(row=row_idx, column=3))
        ws.cell(row=row_idx, column=4, value=order['price']); style_data(ws.cell(row=row_idx, column=4), '0.00')
        ws.cell(row=row_idx, column=5, value=order['qty']); style_data(ws.cell(row=row_idx, column=5), '#,##0')
        ws.cell(row=row_idx, column=6, value=order['amount']); style_data(ws.cell(row=row_idx, column=6), '#,##0.00')

    # 列宽
    widths = [12, 22, 12, 12, 8, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = 'A2'  # 冻结表头


# ============ Sheet 2: KPI 看板 ============

def build_kpi_sheet(wb: Workbook, orders: list[dict], num_days: int) -> None:
    ws = wb.create_sheet('KPI看板')
    # KPI 卡（顶部）
    ws['A1'] = 'KPI'; style_header(ws['A1'])
    ws['B1'] = '今日'; style_header(ws['B1'])
    ws['C1'] = '本周'; style_header(ws['C1'])
    ws['D1'] = '本月'; style_header(ws['D1'])

    # 这些用公式，引用 数据录入 sheet
    today = date.today()
    today_str = today.strftime('%Y-%m-%d')
    week_start = (today - timedelta(days=today.weekday())).strftime('%Y-%m-%d')
    month_start = today.replace(day=1).strftime('%Y-%m-%d')

    metrics = [
        ('GMV (元)', '数据录入!F:F', '#,##0.00'),
        ('订单数', '数据录入!B:B', '#,##0'),
        ('客单价', None, '#,##0.00'),
    ]

    for row_idx, (name, range_addr, fmt) in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        style_data(ws.cell(row=row_idx, column=1))
        for col_idx, (label, period_start) in enumerate(
            [('today', today_str), ('week', week_start), ('month', month_start)], start=2):
            if name == 'GMV (元)':
                f = (f'=SUMIFS(数据录入!F:F, 数据录入!A:A, ">="&DATE({period_start[:4]},'
                     f'{int(period_start[5:7])},{int(period_start[8:10])}))')
            elif name == '订单数':
                f = (f'=COUNTIFS(数据录入!A:A, ">="&DATE({period_start[:4]},'
                     f'{int(period_start[5:7])},{int(period_start[8:10])}))')
            else:  # 客单价 = GMV / 订单数
                gmv_cell = f'{get_column_letter(col_idx)}2'
                cnt_cell = f'{get_column_letter(col_idx)}3'
                f = f'=IFERROR({gmv_cell}/{cnt_cell}, 0)'
            cell = ws.cell(row=row_idx, column=col_idx, value=f)
            style_data(cell, fmt)

    # 每日 GMV 表（用于画图）
    ws['A6'] = '每日 GMV 趋势'
    ws['A6'].font = Font(name='微软雅黑', size=12, bold=True)
    headers = ['日期', '当日 GMV', '7日移动平均']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col_idx, value=h); style_header(c)

    today = date.today()
    for d in range(num_days):
        day = today - timedelta(days=num_days - d - 1)
        row = 8 + d
        date_str = day.strftime('%Y-%m-%d')
        ws.cell(row=row, column=1, value=day); style_data(ws.cell(row=row, column=1), 'yyyy-mm-dd')
        # 当日 GMV：SUMIFS
        ws.cell(row=row, column=2, value=(
            f'=SUMIFS(数据录入!F:F, 数据录入!A:A, A{row})'))
        style_data(ws.cell(row=row, column=2), '#,##0.00')
        # 7 日移动平均
        if d >= 6:
            ws.cell(row=row, column=3, value=f'=AVERAGE(B{row-6}:B{row})')
        else:
            ws.cell(row=row, column=3, value=f'=AVERAGE(B$8:B{row})')
        style_data(ws.cell(row=row, column=3), '#,##0.00')

    # 加图表
    chart = LineChart()
    chart.title = '每日 GMV + 7 日均线'
    chart.x_axis.title = '日期'
    chart.y_axis.title = 'GMV (元)'
    chart.height = 10
    chart.width = 20
    data = Reference(ws, min_col=2, min_row=7, max_col=3, max_row=7 + num_days)
    cats = Reference(ws, min_col=1, min_row=8, max_col=1, max_row=7 + num_days)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'E7')

    # 列宽
    for col_idx, w in enumerate([14, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ Sheet 3: 类目分析 + ABC 分类 ============

def build_category_sheet(wb: Workbook, orders: list[dict]) -> None:
    ws = wb.create_sheet('类目分析')
    ws['A1'] = '类目销售排名 + ABC 分类（A=贡献前 80%, B=80-95%, C=后 5%）'
    ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)

    headers = ['类目', '销售额', '订单数', '占比', '累计占比', 'ABC 分类']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h); style_header(c)

    for row_idx, cat in enumerate(CATEGORIES, start=4):
        ws.cell(row=row_idx, column=1, value=cat); style_data(ws.cell(row=row_idx, column=1))
        ws.cell(row=row_idx, column=2,
                value=f'=SUMIFS(数据录入!F:F, 数据录入!C:C, A{row_idx})')
        style_data(ws.cell(row=row_idx, column=2), '#,##0.00')
        ws.cell(row=row_idx, column=3,
                value=f'=COUNTIFS(数据录入!C:C, A{row_idx})')
        style_data(ws.cell(row=row_idx, column=3), '#,##0')
        ws.cell(row=row_idx, column=4,
                value=f'=B{row_idx}/SUM(B$4:B${3+len(CATEGORIES)})')
        style_data(ws.cell(row=row_idx, column=4), '0.00%')

    # 累计占比 + ABC 分类（按销售额降序计算累计，简化版）
    total_row = 4 + len(CATEGORIES)
    for row_idx in range(4, total_row):
        # 累计占比 = 当前及之前的 D 列累加
        ws.cell(row=row_idx, column=5,
                value=f'=SUM(D$4:D{row_idx})')
        style_data(ws.cell(row=row_idx, column=5), '0.00%')
        # ABC 分类
        ws.cell(row=row_idx, column=6,
                value=f'=IF(E{row_idx}<=0.8, "A", IF(E{row_idx}<=0.95, "B", "C"))')
        style_data(ws.cell(row=row_idx, column=6))

    # 加柱状图
    chart = BarChart()
    chart.type = 'col'
    chart.title = '类目销售额'
    chart.height = 10
    chart.width = 16
    data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=total_row - 1)
    cats = Reference(ws, min_col=1, min_row=4, max_col=1, max_row=total_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'H3')

    for col_idx, w in enumerate([10, 16, 12, 10, 12, 10], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ Sheet 4: 漏斗 ============

def build_funnel_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet('销售漏斗')
    ws['A1'] = '销售漏斗 — 输入每个环节的人数 / 次数'
    ws['A1'].font = Font(name='微软雅黑', size=12, bold=True)

    headers = ['环节', '人次', '环节转化率', '累计转化率']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col_idx, value=h); style_header(c)

    stages = [
        ('曝光', 100000),
        ('点击', 25000),
        ('加购', 8000),
        ('下单', 3500),
        ('支付', 2800),
    ]

    for row_idx, (stage, value) in enumerate(stages, start=4):
        ws.cell(row=row_idx, column=1, value=stage); style_data(ws.cell(row=row_idx, column=1))
        ws.cell(row=row_idx, column=2, value=value); style_data(ws.cell(row=row_idx, column=2), '#,##0')
        # 环节转化率：本环节 / 上环节
        if row_idx == 4:
            ws.cell(row=row_idx, column=3, value=1)  # 第一环节默认 100%
        else:
            ws.cell(row=row_idx, column=3, value=f'=B{row_idx}/B{row_idx-1}')
        style_data(ws.cell(row=row_idx, column=3), '0.00%')
        # 累计转化率：本环节 / 第一环节
        ws.cell(row=row_idx, column=4, value=f'=B{row_idx}/B$4')
        style_data(ws.cell(row=row_idx, column=4), '0.00%')

    # 漏斗图（用条形图模拟）
    chart = BarChart()
    chart.type = 'bar'
    chart.title = '销售漏斗'
    chart.height = 10
    chart.width = 16
    data = Reference(ws, min_col=2, min_row=3, max_col=2, max_row=3 + len(stages))
    cats = Reference(ws, min_col=1, min_row=4, max_col=1, max_row=3 + len(stages))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, 'F3')

    for col_idx, w in enumerate([12, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============ Main ============

def main() -> int:
    parser = argparse.ArgumentParser(description='Generate e-commerce GMV dashboard template')
    parser.add_argument('--output', '-o', default='gmv-dashboard.xlsx')
    parser.add_argument('--days', type=int, default=30, help='How many days of sample data (default 30)')
    args = parser.parse_args()

    out = Path(args.output)
    if out.exists():
        out.unlink()

    wb = Workbook()
    # 删除默认 Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    print(f'[1/4] 生成 {args.days} 天样本订单...')
    orders = make_orders(args.days)
    print(f'      {len(orders)} 条订单')

    print('[2/4] 数据录入 sheet...')
    build_data_sheet(wb, orders)

    print('[3/4] KPI 看板 sheet...')
    build_kpi_sheet(wb, orders, args.days)

    print('[4/4] 类目分析 + 销售漏斗 sheet...')
    build_category_sheet(wb, orders)
    build_funnel_sheet(wb)

    wb.save(out)

    print()
    msg = f'[OK] 模板已生成：{out.absolute()}'
    try:
        print(msg)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((msg + '\n').encode('utf-8'))
    print()
    print('注意：openpyxl 写入的公式在文件层面只是字符串。')
    print('用 Excel 打开会自动重算；或者跑 python scripts/recalc.py 让 LibreOffice 重算。')
    return 0


if __name__ == '__main__':
    sys.exit(main())
