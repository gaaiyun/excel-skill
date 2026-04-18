#!/usr/bin/env python3
"""
dau_mau_cohort.py — 互联网产品留存与活跃模板

生成包含以下 sheet 的模板：
- Sheet 1: 用户日活数据（每日 DAU/新增/流失）
- Sheet 2: 留存矩阵（Cohort 留存 D1/D7/D30）
- Sheet 3: DAU/MAU 比值 + 粘性分析
- Sheet 4: LTV / CAC / 回本周期

用法:
    python scripts/generate_internet/dau_mau_cohort.py --output dau.xlsx
    python scripts/generate_internet/dau_mau_cohort.py --days 90
"""
from __future__ import annotations

import argparse
import random
import sys
from datetime import date, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print('[error] openpyxl not installed. Run: pip install openpyxl', file=sys.stderr)
    sys.exit(2)


HEADER_FONT = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='4472C4')  # 互联网蓝
NORMAL_FONT = Font(name='微软雅黑', size=10)
TOTAL_FONT = Font(name='微软雅黑', size=10, bold=True)
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')


def col(n: int) -> str:
    return get_column_letter(n)


def style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')


def build_daily(wb: Workbook, days: int) -> None:
    ws = wb.create_sheet('1.每日活跃')
    ws['A1'] = f'每日活跃用户数据（最近 {days} 天）'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')

    headers = ['日期', '新增用户', '流失用户', 'DAU', 'WAU 估算', 'MAU 估算', 'DAU/MAU']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    today = date.today()
    base_dau = 100_000
    random.seed(42)

    for i in range(days):
        d = today - timedelta(days=days - 1 - i)
        row = 4 + i
        ws.cell(row=row, column=1, value=d).number_format = 'yyyy-mm-dd'

        new_users = random.randint(1500, 5000)
        churn = random.randint(800, 3000)
        dau = base_dau + (new_users - churn) * 0.6 + random.randint(-3000, 3000)
        base_dau = max(50_000, dau)

        c = ws.cell(row=row, column=2, value=new_users)
        c.fill = INPUT_FILL
        c.number_format = '#,##0'
        c = ws.cell(row=row, column=3, value=churn)
        c.fill = INPUT_FILL
        c.number_format = '#,##0'
        c = ws.cell(row=row, column=4, value=int(dau))
        c.fill = INPUT_FILL
        c.number_format = '#,##0'

        # WAU 7 日均值
        if i >= 6:
            f = f'=AVERAGE(D{row - 6}:D{row})*7'
        else:
            f = f'=AVERAGE(D$4:D{row})*7'
        c = ws.cell(row=row, column=5, value=f)
        c.number_format = '#,##0'

        # MAU 30 日均值
        if i >= 29:
            f = f'=AVERAGE(D{row - 29}:D{row})*30'
        else:
            f = f'=AVERAGE(D$4:D{row})*30'
        c = ws.cell(row=row, column=6, value=f)
        c.number_format = '#,##0'

        # DAU/MAU
        c = ws.cell(row=row, column=7, value=f'=IFERROR(D{row}/F{row},"")')
        c.number_format = '0.0%'

    # 添加 DAU 趋势图
    chart = LineChart()
    chart.title = 'DAU 趋势'
    chart.y_axis.title = 'DAU'
    chart.x_axis.title = '日期'
    data = Reference(ws, min_col=4, min_row=3, max_col=4, max_row=3 + days)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + days)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, f'I3')

    for j in range(1, 8):
        ws.column_dimensions[col(j)].width = 12


def build_cohort(wb: Workbook) -> None:
    ws = wb.create_sheet('2.留存矩阵')
    ws['A1'] = '留存 Cohort 矩阵（按周分组）'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')

    note = (
        '操作说明：\n'
        '行 = 用户首次注册周（cohort）；列 = 注册后第 N 周\n'
        '单元格值 = 该 cohort 在第 N 周仍活跃的用户比例\n'
        '颜色越深表示留存越好（条件格式自动）'
    )
    ws['A2'] = note
    ws['A2'].font = Font(name='微软雅黑', size=9, italic=True, color='808080')
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells('A2:I3')

    headers = ['Cohort 周', '新增'] + [f'W{i}' for i in range(13)]
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=5, column=j, value=h))

    today = date.today()
    random.seed(42)

    n_cohorts = 12
    for i in range(n_cohorts):
        cohort_start = today - timedelta(weeks=n_cohorts - i)
        row = 6 + i
        ws.cell(row=row, column=1,
                value=cohort_start.strftime('%Y-W%V')).font = NORMAL_FONT

        new_users = random.randint(5000, 15000)
        ws.cell(row=row, column=2, value=new_users).number_format = '#,##0'

        # W0-W12 留存率
        for w in range(min(13, n_cohorts - i + 1)):
            if w == 0:
                retention = 1.0
            else:
                # 模拟留存衰减
                retention = max(0.05, 1.0 * (0.7 ** (w * 0.5)) + random.uniform(-0.05, 0.05))
            c = ws.cell(row=row, column=3 + w, value=retention)
            c.number_format = '0.0%'
            c.alignment = Alignment(horizontal='center')
            c.fill = INPUT_FILL

    # 颜色刻度：越浅越红，越深越绿
    ws.conditional_formatting.add(
        f'C6:O{5 + n_cohorts}',
        ColorScaleRule(
            start_type='min', start_color='FFC7CE',
            mid_type='percentile', mid_value=50, mid_color='FFEB9C',
            end_type='max', end_color='C6EFCE',
        )
    )

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    for j in range(3, 16):
        ws.column_dimensions[col(j)].width = 8


def build_stickiness(wb: Workbook) -> None:
    ws = wb.create_sheet('3.粘性分析')
    ws['A1'] = '粘性指标 Dashboard'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')

    headers = ['指标', '值', '说明']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    rows = [
        ('平均 DAU', f"=AVERAGE('1.每日活跃'!D:D)", '7 日 / 30 日 / 全期均可改'),
        ('平均 MAU', f"=AVERAGE('1.每日活跃'!F:F)", '30 日累积去重'),
        ('DAU/MAU 比', '=B4/B5', '高于 20% = 高粘性'),
        ('日均新增', f"=AVERAGE('1.每日活跃'!B:B)", '下降说明获客有问题'),
        ('日均流失', f"=AVERAGE('1.每日活跃'!C:C)", '上升说明产品/竞争问题'),
        ('净增长', '=B7-B8', '负数说明产品在缩水'),
    ]

    for i, (name, formula, desc) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=name).font = TOTAL_FONT
        c = ws.cell(row=i, column=2, value=formula)
        if name == 'DAU/MAU 比':
            c.number_format = '0.0%'
        else:
            c.number_format = '#,##0'
        c.alignment = Alignment(horizontal='right')
        ws.cell(row=i, column=3, value=desc).font = Font(
            name='微软雅黑', size=9, italic=True, color='808080')

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 30


def build_ltv_cac(wb: Workbook) -> None:
    ws = wb.create_sheet('4.LTV_CAC')
    ws['A1'] = 'LTV / CAC / 回本周期'
    ws['A1'].font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')

    headers = ['指标', '值', '公式 / 说明']
    for j, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=j, value=h))

    rows = [
        ('ARPU 月', 30, '每用户月均收入（元）'),
        ('毛利率', 0.7, '70%'),
        ('月留存率', 0.85, '月度留存（用 Cohort 算）'),
        ('月流失率', '=1-B6', '1 - 留存率'),
        ('用户生命周期 (月)', '=1/B7', '1 / 流失率'),
        ('LTV', '=B4*B5*B8', 'ARPU × 毛利率 × 生命周期'),
        ('CAC', 200, '获客成本（元）'),
        ('LTV/CAC 倍数', '=B9/B10', '健康 ≥ 3'),
        ('回本周期 (月)', '=B10/(B4*B5)', 'CAC / (ARPU × 毛利率)'),
    ]

    for i, (name, val, desc) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=name).font = TOTAL_FONT
        c = ws.cell(row=i, column=2, value=val)
        if isinstance(val, (int, float)):
            c.fill = INPUT_FILL
            if name == '毛利率' or name == '月留存率':
                c.number_format = '0.0%'
            else:
                c.number_format = '#,##0.00'
        else:
            if '率' in name:
                c.number_format = '0.0%'
            elif '倍数' in name:
                c.number_format = '0.00'
            else:
                c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal='right')
        ws.cell(row=i, column=3, value=desc).font = Font(
            name='微软雅黑', size=9, italic=True, color='808080')

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 30

    # 高亮 LTV/CAC
    ws['B11'].font = Font(name='微软雅黑', size=11, bold=True, color='C00000')


def main() -> int:
    parser = argparse.ArgumentParser(description='Generate internet/SaaS DAU/MAU/Cohort template')
    parser.add_argument('--output', '-o', default='dau_mau_cohort.xlsx')
    parser.add_argument('--days', type=int, default=60)
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)

    build_daily(wb, args.days)
    build_cohort(wb)
    build_stickiness(wb)
    build_ltv_cac(wb)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f'[ok] Generated: {out}')
    print(f'     Sheets: 1.每日活跃, 2.留存矩阵, 3.粘性分析, 4.LTV_CAC')
    print(f'     Days:   {args.days}')
    print()
    print('Note: Cohort 留存矩阵带颜色刻度；LTV_CAC 输入区为浅黄。')
    return 0


if __name__ == '__main__':
    sys.exit(main())
