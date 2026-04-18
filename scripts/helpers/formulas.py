"""Excel 公式生成 helpers — 把高频公式封装成函数，减少手写错误。"""
from __future__ import annotations

from openpyxl.utils import get_column_letter


def sumifs_range(sum_range: str, criteria_pairs: list[tuple[str, str]]) -> str:
    """生成 SUMIFS 公式字符串。

    Example:
        sumifs_range('Sales!D:D', [('Sales!A:A', '"=Beijing"'), ('Sales!B:B', 'A2')])
        → '=SUMIFS(Sales!D:D,Sales!A:A,"=Beijing",Sales!B:B,A2)'

    Args:
        sum_range: 求和列范围，如 'Sales!D:D'
        criteria_pairs: [(criteria_range, criteria), ...]
    """
    parts = [sum_range]
    for r, c in criteria_pairs:
        parts.append(r)
        parts.append(c)
    return f'=SUMIFS({",".join(parts)})'


def countifs_range(criteria_pairs: list[tuple[str, str]]) -> str:
    """生成 COUNTIFS 公式。"""
    parts = []
    for r, c in criteria_pairs:
        parts.append(r)
        parts.append(c)
    return f'=COUNTIFS({",".join(parts)})'


def growth_rate_yoy(current_cell: str, prev_year_cell: str) -> str:
    """同比增长率（year-on-year）。

    防 0 除：用 IFERROR / IF 包一层
    """
    return f'=IFERROR(({current_cell}-{prev_year_cell})/ABS({prev_year_cell}),"")'


def growth_rate_mom(current_cell: str, prev_month_cell: str) -> str:
    """环比增长率（month-on-month）。"""
    return growth_rate_yoy(current_cell, prev_month_cell)


def safe_division(numerator: str, denominator: str, fallback: str = '0') -> str:
    """除法防 0：分母为 0 时返回 fallback。

    Example:
        safe_division('B2', 'C2')                  → '=IF(C2=0,0,B2/C2)'
        safe_division('B2', 'C2', fallback='""')   → '=IF(C2=0,"",B2/C2)'
    """
    return f'=IF({denominator}=0,{fallback},{numerator}/{denominator})'


def cumulative_sum_range(start_cell: str, current_cell: str) -> str:
    """累计求和：从 start_cell 到 current_cell 之间的所有值之和。

    Example:
        cumulative_sum_range('B$2', 'B2') → '=SUM(B$2:B2)'

    用法：把这个公式放在第 2 行后下拉，会自动累加。
    """
    return f'=SUM({start_cell}:{current_cell})'


def vlookup(lookup_value: str, table_array: str, col_index: int,
            exact: bool = True) -> str:
    """VLOOKUP 公式。exact=True 表示精确匹配（最常用）。"""
    range_lookup = 'FALSE' if exact else 'TRUE'
    return f'=VLOOKUP({lookup_value},{table_array},{col_index},{range_lookup})'


def xlookup(lookup_value: str, lookup_array: str, return_array: str,
            if_not_found: str = '""') -> str:
    """XLOOKUP（Excel 365+，比 VLOOKUP 更现代）。

    Example:
        xlookup('A2', 'Products!A:A', 'Products!B:B', '"未找到"')
    """
    return f'=XLOOKUP({lookup_value},{lookup_array},{return_array},{if_not_found})'


def index_match(lookup_value: str, return_array: str, match_array: str,
                match_type: int = 0) -> str:
    """INDEX/MATCH 组合（XLOOKUP 不可用时的经典替代）。

    match_type: 0=精确，1=升序近似，-1=降序近似
    """
    return f'=INDEX({return_array},MATCH({lookup_value},{match_array},{match_type}))'


def cagr(start_value: str, end_value: str, n_years: float) -> str:
    """复合年增长率 (Compound Annual Growth Rate)。

    Example:
        cagr('B2', 'B6', 4)  # 4 年 CAGR
    """
    return f'=({end_value}/{start_value})^(1/{n_years})-1'


def npv(rate_cell: str, cash_flow_range: str) -> str:
    """净现值 NPV。注意 Excel 的 NPV 假设第一个值是第 1 期末，不是第 0 期。"""
    return f'=NPV({rate_cell},{cash_flow_range})'


def irr(cash_flow_range: str, guess: str = '0.1') -> str:
    """内部收益率 IRR。"""
    return f'=IRR({cash_flow_range},{guess})'


def col_letter_range(col_start: int, col_end: int) -> str:
    """列号转字母范围，如 (1, 3) → 'A:C'。"""
    return f'{get_column_letter(col_start)}:{get_column_letter(col_end)}'
