"""通用 Excel 样式 helpers — 跨 4 大行业模板复用。"""
from __future__ import annotations

from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


# ============ 品牌色定义（4 大行业可调） ============

BRAND_COLORS = {
    'finance': {
        'header_bg': '1F4E78',     # 深蓝
        'header_fg': 'FFFFFF',
        'kpi_bg': 'D9E1F2',        # 浅蓝
        'kpi_accent': '1F4E78',
        'positive': '00B050',       # 绿
        'negative': 'C00000',       # 红
    },
    'fmcg': {
        'header_bg': '2E75B6',     # 中蓝
        'header_fg': 'FFFFFF',
        'kpi_bg': 'DEEBF7',
        'kpi_accent': '2E75B6',
        'positive': '70AD47',
        'negative': 'C55A11',
    },
    'ecommerce': {
        'header_bg': 'C00000',     # 电商红
        'header_fg': 'FFFFFF',
        'kpi_bg': 'FCE4D6',
        'kpi_accent': 'C00000',
        'positive': '00B050',
        'negative': '7030A0',
    },
    'internet': {
        'header_bg': '4472C4',     # 互联网蓝
        'header_fg': 'FFFFFF',
        'kpi_bg': 'E7E6E6',
        'kpi_accent': '4472C4',
        'positive': '70AD47',
        'negative': 'C00000',
    },
}


def apply_header_style(ws, row: int, col_start: int, col_end: int,
                       industry: str = 'finance') -> None:
    """套用表头样式（粗体白字 + 行业色背景 + 居中 + 边框）。"""
    colors = BRAND_COLORS.get(industry, BRAND_COLORS['finance'])
    fill = PatternFill('solid', fgColor=colors['header_bg'])
    font = Font(bold=True, color=colors['header_fg'], size=11)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='thin', color=colors['header_bg']),
        bottom=Side(style='thin', color=colors['header_bg']),
    )
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[row].height = 28


def apply_kpi_style(ws, cell_ref: str, industry: str = 'finance') -> None:
    """KPI 大数字样式：行业 accent 色背景 + 大字 + 居中。"""
    colors = BRAND_COLORS.get(industry, BRAND_COLORS['finance'])
    cell = ws[cell_ref]
    cell.fill = PatternFill('solid', fgColor=colors['kpi_bg'])
    cell.font = Font(bold=True, size=18, color=colors['kpi_accent'])
    cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_money_format(ws, range_ref: str, currency: str = 'CNY',
                       decimals: int = 2) -> None:
    """金额格式：千分位 + 小数 + 货币符号。

    currency: 'CNY' / 'USD' / 'EUR' / 'JPY'
    """
    fmt_map = {
        'CNY': f'¥#,##0.{"0" * decimals};[Red]-¥#,##0.{"0" * decimals}',
        'USD': f'$#,##0.{"0" * decimals};[Red]-$#,##0.{"0" * decimals}',
        'EUR': f'€#,##0.{"0" * decimals};[Red]-€#,##0.{"0" * decimals}',
        'JPY': '¥#,##0;[Red]-¥#,##0',  # 日元无小数
        'plain': f'#,##0.{"0" * decimals};[Red]-#,##0.{"0" * decimals}',
    }
    fmt = fmt_map.get(currency, fmt_map['plain'])
    for row in ws[range_ref]:
        for cell in row:
            cell.number_format = fmt


def apply_percent_format(ws, range_ref: str, decimals: int = 2,
                         show_sign: bool = False) -> None:
    """百分比格式。

    decimals: 小数位
    show_sign: 是否显示 + 号（YoY 增长率常用）
    """
    if show_sign:
        fmt = f'+0.{"0" * decimals}%;-0.{"0" * decimals}%;0.{"0" * decimals}%'
    else:
        fmt = f'0.{"0" * decimals}%'
    for row in ws[range_ref]:
        for cell in row:
            cell.number_format = fmt


def apply_date_format(ws, range_ref: str, fmt: str = 'yyyy-mm-dd') -> None:
    """日期格式：默认 ISO，可改 'yyyy/m/d' / 'yyyy年m月' / 'mmm-yy' 等。"""
    for row in ws[range_ref]:
        for cell in row:
            cell.number_format = fmt


def apply_thin_border(ws, range_ref: str, color: str = 'BFBFBF') -> None:
    """加细边框（默认浅灰）。"""
    side = Side(style='thin', color=color)
    border = Border(left=side, right=side, top=side, bottom=side)
    for row in ws[range_ref]:
        for cell in row:
            cell.border = border


def apply_zebra_stripes(ws, range_ref: str, color: str = 'F2F2F2') -> None:
    """斑马纹（隔行底色），提升大表可读性。"""
    fill = PatternFill('solid', fgColor=color)
    for i, row in enumerate(ws[range_ref]):
        if i % 2 == 1:  # 偶数行（0-based 的奇数 index）
            for cell in row:
                cell.fill = fill


def auto_width(ws, max_col: int = None, min_width: float = 8.0,
               max_width: float = 60.0, padding: float = 2.0) -> None:
    """根据每列内容自动调整列宽。"""
    if max_col is None:
        max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = min_width
        for cell in ws[col_letter]:
            try:
                val = str(cell.value) if cell.value is not None else ''
                # 中文字符算两个宽度
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, width)
            except Exception:
                continue
        ws.column_dimensions[col_letter].width = min(max_len + padding, max_width)


def freeze_top_row(ws, freeze_at: str = 'A2') -> None:
    """冻结顶行（默认第 1 行）。"""
    ws.freeze_panes = freeze_at


def freeze_top_left(ws, freeze_at: str = 'B2') -> None:
    """冻结顶行 + 首列（左上角十字）。"""
    ws.freeze_panes = freeze_at
